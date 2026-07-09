"""
Topic LinkedIn #3 -- "Markowitz applique a 8 actions : la frontiere efficiente algerienne"
Bourse d'Alger (BAA) : optimisation de portefeuille moyenne-variance.

Ce script lit directement le classeur Excel BAA_Prix_Dividendes_Portefeuille_v6.xlsx
(feuilles "Statistiques", "Corrélation" et "Inflation") au lieu de donnees codees en dur.

NOTE METHODOLOGIQUE : cette optimisation est reconstruite a partir des statistiques
resumees (rendement/vol annualises + correlations), pas des rendements journaliers bruts.
Les poids obtenus peuvent donc differer legerement de ceux calcules sur donnees completes
(portefeuilles de reference dans la feuille "Portefeuilles" du classeur).

Necessite : openpyxl, numpy, scipy, matplotlib
Auteur : Younes Benzouai
"""

import sys
import os
import re
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from scipy.optimize import minimize

DEFAULT_EXCEL_PATH = r"C:\Users\youne\OneDrive\Desktop\LinkedIn\DZ\Outil d'optimisation de portefeuille boursier\BAA_Prix_Dividendes_Portefeuille_v6.xlsx"


def _resolve_excel_path(default_path):
    """Priorite : argument CLI > variable d'env BAA_XLSX_PATH > fichier .xlsx present
    a cote de ce script (cas du depot GitHub) > chemin par defaut. Ignore les arguments
    injectes par Jupyter/IPython (ex: '-f ...kernel.json')."""
    is_notebook = "ipykernel" in sys.argv[0] or "ipykernel" in sys.modules
    if not is_notebook and len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        return sys.argv[1]
    env_path = os.environ.get("BAA_XLSX_PATH")
    if env_path:
        return env_path
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        script_dir = os.getcwd()
    local_path = os.path.join(script_dir, "BAA_Prix_Dividendes_Portefeuille_v6.xlsx")
    if os.path.exists(local_path):
        return local_path
    return default_path


EXCEL_PATH = _resolve_excel_path(DEFAULT_EXCEL_PATH)


def _open(path):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fichier introuvable : {path}\n"
            "-> Verifie DEFAULT_EXCEL_PATH en haut du script, ou definis la variable\n"
            "   d'environnement BAA_XLSX_PATH avant de lancer le script/la cellule."
        )
    return openpyxl.load_workbook(path, data_only=True)


def load_statistiques(path, sheet_name="Statistiques"):
    """Charge le tableau 'Indicateur x Ticker' de la feuille Statistiques.
    Retourne (tickers, names, {indicateur: {ticker: valeur}})."""
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Indicateur":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError("Ligne d'en-tete 'Indicateur' introuvable dans la feuille Statistiques.")

    header_cells = ws[header_row]
    tickers, names = [], {}
    for c in header_cells[1:]:
        if c.value is None:
            break
        parts = str(c.value).split("\n")
        t = parts[0].strip()
        tickers.append(t)
        names[t] = parts[1].strip() if len(parts) > 1 else t

    indicators = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        label = row[0].value
        if not label or not isinstance(label, str):
            continue
        values = [c.value for c in row[1:1 + len(tickers)]]
        if all(v is None for v in values):
            continue
        indicators[label.strip()] = dict(zip(tickers, values))

    return tickers, names, indicators


def load_correlation(path, tickers, sheet_name="Corrélation"):
    """Charge la matrice de correlation en respectant l'ordre des tickers fourni
    (celui de la feuille Statistiques), en verifiant que les deux feuilles concordent."""
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and "Corr" in str(row[0].value):
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Ligne d'en-tete introuvable dans la feuille {sheet_name}.")

    header_cells = ws[header_row]
    sheet_tickers = [str(c.value).split("\n")[0].strip() for c in header_cells[1:1 + len(tickers)]]
    if sheet_tickers != tickers:
        raise ValueError(
            f"Ordre des tickers different entre Statistiques {tickers} et {sheet_name} {sheet_tickers}."
        )

    n = len(tickers)
    corr = np.zeros((n, n))
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=header_row + n)):
        for j in range(n):
            corr[i, j] = row[1 + j].value
    return corr


def load_rf(path, sheet_name="Inflation"):
    wb = _open(path)
    ws = wb[sheet_name]
    pct_pattern = re.compile(r"^-?\d+([.,]\d+)?%$")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label, value = row[0].value, row[1].value if len(row) > 1 else None
        if isinstance(label, str) and label.lower().startswith("taux sans risque nominal") \
                and isinstance(value, str) and pct_pattern.match(value.strip()):
            return float(value.strip().replace("%", "").replace(",", ".")) / 100
    return 0.03  # valeur par defaut si non trouve


tickers, names, stats = load_statistiques(EXCEL_PATH)
ann_return = np.array([stats["Rendement Annualisé (%)"][t] for t in tickers])
ann_vol = np.array([stats["Volatilité Annualisée (%)"][t] for t in tickers])
corr = load_correlation(EXCEL_PATH, tickers)
RF = load_rf(EXCEL_PATH)

print(f"Source : {EXCEL_PATH}")
print(f"Rf utilise : {RF*100:.2f}%")
print(f"Univers : {tickers}\n")

cov = np.outer(ann_vol, ann_vol) * corr
n = len(tickers)


def perf(w):
    ret = w @ ann_return
    vol = np.sqrt(w @ cov @ w)
    sharpe = (ret - RF) / vol if vol > 0 else 0
    return ret, vol, sharpe


def neg_sharpe(w):
    return -perf(w)[2]


def portfolio_vol(w):
    return perf(w)[1]


bounds = tuple((0, 1) for _ in range(n))
constraints = {"type": "eq", "fun": lambda w: np.sum(w) - 1}
w0 = np.repeat(1 / n, n)

max_sharpe = minimize(neg_sharpe, w0, method="SLSQP", bounds=bounds, constraints=constraints)
min_vol = minimize(portfolio_vol, w0, method="SLSQP", bounds=bounds, constraints=constraints)

print("=== Portefeuille Max Sharpe (calcule) ===")
for t, w in zip(tickers, max_sharpe.x):
    if w > 0.005:
        print(f"  {t:4s} {w*100:5.1f}%")
r, v, s = perf(max_sharpe.x)
print(f"  -> Rendement {r*100:.2f}% | Volatilite {v*100:.2f}% | Sharpe {s:.2f}\n")

print("=== Portefeuille Min Volatilite (calcule) ===")
for t, w in zip(tickers, min_vol.x):
    if w > 0.005:
        print(f"  {t:4s} {w*100:5.1f}%")
r, v, s = perf(min_vol.x)
print(f"  -> Rendement {r*100:.2f}% | Volatilite {v*100:.2f}% | Sharpe {s:.2f}\n")

r_eq, v_eq, s_eq = perf(w0)
print(f"Equi-pondere : Rendement {r_eq*100:.2f}% | Volatilite {v_eq*100:.2f}% | Sharpe {s_eq:.2f}\n")

np.random.seed(42)
N_PORT = 20000
results = np.zeros((3, N_PORT))
for i in range(N_PORT):
    w = np.random.dirichlet(np.ones(n))
    r, v, s = perf(w)
    results[:, i] = [r, v, s]

fig, ax = plt.subplots(figsize=(10, 7))
sc = ax.scatter(results[1] * 100, results[0] * 100, c=results[2], cmap="viridis", s=6, alpha=0.5)
plt.colorbar(sc, label="Ratio de Sharpe")

r_ms, v_ms, _ = perf(max_sharpe.x)
r_mv, v_mv, _ = perf(min_vol.x)
ax.scatter([v_ms * 100], [r_ms * 100], marker="*", s=350, color="#C00000", edgecolor="black", label="Max Sharpe", zorder=5)
ax.scatter([v_mv * 100], [r_mv * 100], marker="*", s=350, color="#1F3864", edgecolor="black", label="Min Volatilite", zorder=5)
ax.scatter([v_eq * 100], [r_eq * 100], marker="D", s=90, color="orange", edgecolor="black", label="Equi-pondere", zorder=5)

for i, t in enumerate(tickers):
    ax.scatter(ann_vol[i] * 100, ann_return[i] * 100, marker="x", s=80, color="grey")
    ax.annotate(t, (ann_vol[i] * 100, ann_return[i] * 100), fontsize=8, xytext=(4, 4), textcoords="offset points")

ax.set_xlabel("Volatilite annualisee (%)")
ax.set_ylabel("Rendement annualise (%)")
ax.set_title(f"Frontiere efficiente -- {n} actions de la Bourse d'Alger\n(20 000 portefeuilles simules, Rf = {RF*100:.1f}%)", fontsize=12)
ax.legend(fontsize=9)
ax.spines[["top", "right"]].set_visible(False)

plt.tight_layout()
plt.savefig("03_optimisation_portefeuille.png", dpi=150)
print("Graphique sauvegarde : 03_optimisation_portefeuille.png")
