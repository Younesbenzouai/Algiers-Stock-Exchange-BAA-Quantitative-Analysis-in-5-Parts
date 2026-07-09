"""
Topic LinkedIn #1 -- "Un marche de 5,6 Mds$ domine par deux banques"
Bourse d'Alger (BAA) : structure et concentration du marche, juillet 2026.

Ce script lit directement le classeur Excel BAA_Prix_Dividendes_Portefeuille_v6.xlsx
(feuille "Capitalisation") au lieu de donnees codees en dur : modifie le fichier source
et relance le script (ou la cellule, si tu es dans Jupyter) pour mettre a jour le post/graphique.

Necessite : openpyxl, matplotlib, numpy
Auteur : Younes Benzouai
"""

import sys
import os
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# --- Chemin du classeur source -------------------------------------------------
DEFAULT_EXCEL_PATH = r"C:\Users\youne\OneDrive\Desktop\LinkedIn\DZ\Outil d'optimisation de portefeuille boursier\BAA_Prix_Dividendes_Portefeuille_v6.xlsx"


def _resolve_excel_path(default_path):
    """Determine le chemin du fichier Excel a utiliser.

    Priorite : argument de ligne de commande > variable d'environnement BAA_XLSX_PATH >
    chemin par defaut ci-dessus. Ignore automatiquement les arguments injectes par
    Jupyter/IPython (ex: '-f ...kernel.json'), pour que le script fonctionne aussi bien
    lance en notebook qu'en ligne de commande.
    """
    is_notebook = "ipykernel" in sys.argv[0] or "ipykernel" in sys.modules
    if not is_notebook and len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        return sys.argv[1]
    return os.environ.get("BAA_XLSX_PATH", default_path)


EXCEL_PATH = _resolve_excel_path(DEFAULT_EXCEL_PATH)
SHEET_NAME = "Capitalisation"
USD_DZD = 133.1054  # taux de change utilise pour la conversion en Mds USD (a ajuster si besoin)


def load_market_cap(path, sheet_name=SHEET_NAME):
    """Lit le tableau 'RESUME - DERNIERE DATE DISPONIBLE' de la feuille Capitalisation.

    Repere automatiquement la ligne d'en-tete (colonne A = 'Ticker') puis lit les lignes
    suivantes jusqu'a la ligne vide ou la ligne 'TOTAL BAA'. Ne depend donc pas d'un
    numero de ligne fixe : le script reste valide si le classeur est mis a jour.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fichier introuvable : {path}\n"
            "-> Verifie le chemin en haut du script (DEFAULT_EXCEL_PATH), ou definis la\n"
            "   variable d'environnement BAA_XLSX_PATH avant de lancer le script/la cellule."
        )

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' absente. Feuilles disponibles : {wb.sheetnames}")
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Ticker":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError("Ligne d'en-tete 'Ticker' introuvable dans la feuille Capitalisation.")

    headers = [c.value for c in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    required = ["Ticker", "Société", "Secteur", "Cap. (Mrd DZD)"]
    missing = [c for c in required if c not in col]
    if missing:
        raise ValueError(f"Colonnes manquantes dans '{sheet_name}' : {missing} (colonnes trouvees : {headers})")

    data = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        ticker = row[col["Ticker"]].value
        if ticker is None:
            break
        if str(ticker).strip().upper().startswith("TOTAL"):
            break
        data[str(ticker).strip()] = {
            "name": row[col["Société"]].value,
            "sector": row[col["Secteur"]].value,
            "cap_mrd": float(row[col["Cap. (Mrd DZD)"]].value),
        }

    if not data:
        raise ValueError("Aucune ligne de donnees lue sous l'en-tete 'Ticker'.")
    return data


market_cap = load_market_cap(EXCEL_PATH)
print(f"Source : {EXCEL_PATH}")
print(f"Feuille : {SHEET_NAME} -- {len(market_cap)} valeurs chargees\n")

total_cap = sum(v["cap_mrd"] for v in market_cap.values())
total_usd_bn = total_cap * 1e9 / USD_DZD / 1e9

print(f"Capitalisation totale BAA : {total_cap:.2f} Mrd DZD (~{total_usd_bn:.2f} Mds USD)")
print(f"Nombre de societes cotees : {len(market_cap)}\n")

weights = {t: v["cap_mrd"] / total_cap for t, v in market_cap.items()}
hhi = sum(w ** 2 for w in weights.values()) * 10_000

print("Poids par valeur :")
for t, w in sorted(weights.items(), key=lambda x: -x[1]):
    print(f"  {t:4s} {str(market_cap[t]['name']):22s} {w*100:5.1f}%")

print(f"\nIndice Herfindahl-Hirschman (HHI) : {hhi:.0f}")
print("  > 2500 = marche hautement concentre (seuil DOJ/FTC)" if hhi > 2500 else "  marche non concentre")

sector_weights = {}
for t, v in market_cap.items():
    sector_weights[v["sector"]] = sector_weights.get(v["sector"], 0) + weights[t]

print("\nPoids par secteur :")
for s, w in sorted(sector_weights.items(), key=lambda x: -x[1]):
    print(f"  {s:16s} {w*100:5.1f}%")

tickers_sorted = sorted(market_cap.keys(), key=lambda t: -market_cap[t]["cap_mrd"])
caps = [market_cap[t]["cap_mrd"] for t in tickers_sorted]
labels = [f"{t}\n{market_cap[t]['name']}" for t in tickers_sorted]
colors = ["#1F3864" if market_cap[t]["sector"] == "Banque" else "#8FAADC" for t in tickers_sorted]

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 6), gridspec_kw={"width_ratios": [1.6, 1]})

bars = ax1.bar(labels, caps, color=colors, edgecolor="white")
ax1.set_ylabel("Capitalisation (Mrd DZD)")
ax1.set_title(f"Bourse d'Alger : capitalisation par valeur\nTotal = {total_cap:.1f} Mrd DZD (~{total_usd_bn:.1f} Mds USD)", fontsize=11)
ax1.tick_params(axis="x", labelsize=8)
for b, c in zip(bars, caps):
    ax1.text(b.get_x() + b.get_width()/2, c + 8, f"{c:.0f}", ha="center", fontsize=8)
ax1.spines[["top", "right"]].set_visible(False)

sectors_sorted = sorted(sector_weights.items(), key=lambda x: -x[1])
ax2.pie(
    [w for _, w in sectors_sorted],
    labels=[f"{s}\n{w*100:.1f}%" for s, w in sectors_sorted],
    colors=plt.cm.Blues(np.linspace(0.9, 0.3, len(sectors_sorted))),
    startangle=90,
    textprops={"fontsize": 8},
)
ax2.set_title(f"Repartition sectorielle\nHHI = {hhi:.0f} (marche concentre)", fontsize=11)

plt.tight_layout()
plt.savefig("01_concentration_marche.png", dpi=150)
print("\nGraphique sauvegarde : 01_concentration_marche.png")
