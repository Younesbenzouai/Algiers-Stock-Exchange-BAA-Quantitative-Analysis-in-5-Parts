"""
Topic LinkedIn #4 -- "Correlations a zero : diversification reelle ou illusion de liquidite ?"
Bourse d'Alger (BAA) : matrice de correlation confrontee a la liquidite reelle des titres.

Ce script lit directement le classeur Excel BAA_Prix_Dividendes_Portefeuille_v6.xlsx
(feuilles "Corrélation" et "Volumes BOC") au lieu de donnees codees en dur.

Necessite : openpyxl, numpy, matplotlib
Auteur : Younes Benzouai
"""

import sys
import os
import numpy as np
import openpyxl
import matplotlib.pyplot as plt

DEFAULT_EXCEL_PATH = r"C:\Users\youne\OneDrive\Desktop\LinkedIn\DZ\Outil d'optimisation de portefeuille boursier\BAA_Prix_Dividendes_Portefeuille_v6.xlsx"


def _resolve_excel_path(default_path):
    """Priorite : argument CLI > variable d'env BAA_XLSX_PATH > chemin par defaut.
    Ignore les arguments injectes par Jupyter/IPython (ex: '-f ...kernel.json')."""
    is_notebook = "ipykernel" in sys.argv[0] or "ipykernel" in sys.modules
    if not is_notebook and len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        return sys.argv[1]
    return os.environ.get("BAA_XLSX_PATH", default_path)


EXCEL_PATH = _resolve_excel_path(DEFAULT_EXCEL_PATH)


def _open(path):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fichier introuvable : {path}\n"
            "-> Verifie DEFAULT_EXCEL_PATH en haut du script, ou definis la variable\n"
            "   d'environnement BAA_XLSX_PATH avant de lancer le script/la cellule."
        )
    return openpyxl.load_workbook(path, data_only=True)


def load_correlation(path, sheet_name="Corrélation"):
    """Charge la matrice de correlation complete + la liste des tickers (ordre de la feuille)."""
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and "Corr" in str(row[0].value):
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Ligne d'en-tete introuvable dans la feuille {sheet_name}.")

    header_cells = ws[header_row]
    tickers = []
    for c in header_cells[1:]:
        if c.value is None:
            break
        tickers.append(str(c.value).split("\n")[0].strip())

    n = len(tickers)
    corr = np.zeros((n, n))
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=header_row + n)):
        for j in range(n):
            corr[i, j] = row[1 + j].value
    return tickers, corr


def load_liquidity(path, tickers, sheet_name="Volumes BOC"):
    """Charge le '% Seances avec tx' par ticker depuis la feuille Volumes BOC."""
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Ticker":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Ligne d'en-tete 'Ticker' introuvable dans la feuille {sheet_name}.")

    headers = [c.value for c in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}
    if "% Séances avec tx" not in col:
        raise ValueError(f"Colonne '% Séances avec tx' introuvable. Colonnes : {headers}")

    pct_active = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        ticker = row[col["Ticker"]].value
        if ticker is None:
            break
        pct_active[str(ticker).strip()] = float(row[col["% Séances avec tx"]].value or 0)

    return {t: pct_active.get(t, 0.0) for t in tickers}


tickers, corr = load_correlation(EXCEL_PATH)
pct_active = load_liquidity(EXCEL_PATH, tickers)
n = len(tickers)

print(f"Source : {EXCEL_PATH}")
print(f"Univers : {tickers}\n")

mean_abs_corr = {}
for i, t in enumerate(tickers):
    others = [abs(corr[i, j]) for j in range(n) if j != i]
    mean_abs_corr[t] = np.mean(others)

print(f"{'Ticker':6s} {'% seances actives':>18s} {'|Correlation| moy.':>20s}")
for t in tickers:
    print(f"{t:6s} {pct_active[t]*100:17.1f}% {mean_abs_corr[t]:19.3f}")

corr_liquidity = np.corrcoef(
    [pct_active[t] for t in tickers],
    [mean_abs_corr[t] for t in tickers],
)[0, 1]
print(f"\nCorrelation (liquidite, |correlation moyenne|) entre titres = {corr_liquidity:+.2f}")
print("-> Plus un titre est liquide, plus sa correlation mesuree avec le reste du marche est elevee :")
print("   signe que les correlations quasi nulles des titres illiquides sont un ARTEFACT statistique,")
print("   pas une vraie diversification economique.")

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

im = ax1.imshow(corr, cmap="RdBu_r", vmin=-1, vmax=1)
ax1.set_xticks(range(n)); ax1.set_xticklabels(tickers, fontsize=9)
ax1.set_yticks(range(n)); ax1.set_yticklabels(tickers, fontsize=9)
for i in range(n):
    for j in range(n):
        ax1.text(j, i, f"{corr[i,j]:.2f}", ha="center", va="center", fontsize=7,
                  color="white" if abs(corr[i, j]) > 0.5 else "black")
ax1.set_title("Matrice de correlation des rendements\n(quasi nulle presque partout)", fontsize=11)
plt.colorbar(im, ax=ax1, fraction=0.046, pad=0.04)

x = [pct_active[t] * 100 for t in tickers]
y = [mean_abs_corr[t] for t in tickers]
ax2.scatter(x, y, s=100, color="#1F3864", zorder=3)
for t in tickers:
    ax2.annotate(t, (pct_active[t] * 100, mean_abs_corr[t]), fontsize=9, xytext=(5, 5), textcoords="offset points")
z = np.polyfit(x, y, 1)
xs = np.linspace(min(x), max(x), 50)
ax2.plot(xs, np.polyval(z, xs), "--", color="#C00000", label=f"Tendance (r = {corr_liquidity:.2f})")
ax2.set_xlabel("% de seances avec transaction (liquidite)")
ax2.set_ylabel("Correlation absolue moyenne avec les autres titres")
ax2.set_title("Illusion de diversification :\ncorrelation vs liquidite", fontsize=11)
ax2.legend(fontsize=9)
ax2.spines[["top", "right"]].set_visible(False)

plt.tight_layout()
plt.savefig("04_correlation_liquidite.png", dpi=150)
print("\nGraphique sauvegarde : 04_correlation_liquidite.png")
