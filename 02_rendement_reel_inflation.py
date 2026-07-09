"""
Topic LinkedIn #2 -- "Le piege du rendement nominal : ce que l'inflation cache vraiment"
Bourse d'Alger (BAA) : rendement nominal vs rendement REEL (formule de Fisher).

Ce script lit directement le classeur Excel BAA_Prix_Dividendes_Portefeuille_v6.xlsx
(feuilles "Statistiques" et "Inflation") au lieu de donnees codees en dur.

Necessite : openpyxl, matplotlib
Auteur : Younes Benzouai
"""

import sys
import os
import re
import openpyxl
import matplotlib.pyplot as plt

DEFAULT_EXCEL_PATH = r"C:\Users\youne\OneDrive\Desktop\LinkedIn\DZ\Outil d'optimisation de portefeuille boursier\BAA_Prix_Dividendes_Portefeuille_v6.xlsx"


def _resolve_excel_path(default_path):
    """Priorite : argument CLI > variable d'env BAA_XLSX_PATH > fichier .xlsx present
    a cote de ce script (cas du depot GitHub) > chemin par defaut. Ignore les arguments
    injectes par Jupyter/IPython (ex: '-f ...kernel.json')."""
    is_notebook = "ipykernel" in sys.argv[0] or "ipykernel" in sys.modules
    if not is_notebook and len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        return sys.argv[1]
    env_path = os.environ.get("BAA_XLSX_PATH")
    if env_path:
        return env_path
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        script_dir = os.getcwd()
    local_path = os.path.join(script_dir, "BAA_Prix_Dividendes_Portefeuille_v6.xlsx")
    if os.path.exists(local_path):
        return local_path
    return default_path


EXCEL_PATH = _resolve_excel_path(DEFAULT_EXCEL_PATH)


def _open(path):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fichier introuvable : {path}\n"
            "-> Verifie DEFAULT_EXCEL_PATH en haut du script, ou definis la variable\n"
            "   d'environnement BAA_XLSX_PATH avant de lancer le script/la cellule."
        )
    return openpyxl.load_workbook(path, data_only=True)


def load_statistiques(path, sheet_name="Statistiques"):
    """Charge le tableau 'Indicateur x Ticker' de la feuille Statistiques.

    Repere la ligne d'en-tete (colonne A = 'Indicateur'), extrait les tickers depuis
    les cellules d'en-tete (format 'TICKER\\nNom Societe'), puis lit chaque ligne
    d'indicateur en ignorant les lignes de separation de section (ex: '-- PRIX --').
    Retourne (tickers, names, {indicateur: {ticker: valeur}}).
    """
    wb = _open(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' absente. Feuilles disponibles : {wb.sheetnames}")
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Indicateur":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError("Ligne d'en-tete 'Indicateur' introuvable dans la feuille Statistiques.")

    header_cells = ws[header_row]
    tickers, names = [], {}
    for c in header_cells[1:]:
        if c.value is None:
            break
        parts = str(c.value).split("\n")
        t = parts[0].strip()
        tickers.append(t)
        names[t] = parts[1].strip() if len(parts) > 1 else t

    indicators = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        label = row[0].value
        if not label or not isinstance(label, str):
            continue
        values = [c.value for c in row[1:1 + len(tickers)]]
        if all(v is None for v in values):
            continue  # ligne de separation de section
        indicators[label.strip()] = dict(zip(tickers, values))

    return tickers, names, indicators


def load_inflation_dashboard(path, sheet_name="Inflation"):
    """Lit les indicateurs macro (Rf nominal, inflation moyenne) de la feuille Inflation.

    Ne garde que les lignes 'Label | Valeur' ou Valeur est un pourcentage NUMERIQUE
    (ex: '3.0%'), pour ignorer les en-tetes de tableau comme 'IPC Global (%)'.
    """
    wb = _open(path)
    ws = wb[sheet_name]
    pct_pattern = re.compile(r"^-?\d+([.,]\d+)?%$")
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label, value = row[0].value, row[1].value if len(row) > 1 else None
        if isinstance(label, str) and isinstance(value, str) and pct_pattern.match(value.strip()):
            out[label.strip()] = float(value.strip().replace("%", "").replace(",", ".")) / 100
    return out


tickers, names, stats = load_statistiques(EXCEL_PATH)
macro = load_inflation_dashboard(EXCEL_PATH)

rf_key = next((k for k in macro if k.lower().startswith("taux sans risque nominal")), None)
infl_key = next((k for k in macro if k.lower().startswith("inflation moyenne")), None)
if rf_key is None or infl_key is None:
    raise ValueError(f"Indicateurs Rf / inflation introuvables dans la feuille Inflation. Cles lues : {list(macro)}")

RF_NOMINAL = macro[rf_key]
INFLATION_PERIODE = macro[infl_key]
rf_reel = (1 + RF_NOMINAL) / (1 + INFLATION_PERIODE) - 1

print(f"Source : {EXCEL_PATH}")
print(f"Taux sans risque nominal      : {RF_NOMINAL*100:+.2f}%")
print(f"Inflation moyenne (periode)   : {INFLATION_PERIODE*100:+.2f}%")
print(f"Taux sans risque REEL (Fisher): {rf_reel*100:+.2f}%  <-- seuil de rentabilite reel\n")

var_prix = stats["Variation Prix (%)"]
div_yield = stats["Rendement Dividendes (%)"]

nominal_total_return = {t: (var_prix[t] or 0) + (div_yield[t] or 0) for t in tickers}
real_return = {t: (1 + r) / (1 + INFLATION_PERIODE) - 1 for t, r in nominal_total_return.items()}

print(f"{'Ticker':6s} {'Nominal':>9s} {'Reel':>9s}  {'Bat le Rf reel ?':>18s}")
for t in sorted(real_return, key=lambda x: -real_return[x]):
    beats = "OUI" if real_return[t] > rf_reel else "non"
    print(f"{t:6s} {nominal_total_return[t]*100:8.1f}% {real_return[t]*100:8.1f}%  {beats:>18s}")

tickers_sorted = sorted(real_return, key=lambda x: -real_return[x])
nominal_vals = [nominal_total_return[t] * 100 for t in tickers_sorted]
real_vals = [real_return[t] * 100 for t in tickers_sorted]

fig, ax = plt.subplots(figsize=(11, 6.5))
x = range(len(tickers_sorted))
w = 0.35

ax.bar([i - w/2 for i in x], nominal_vals, width=w, label="Rendement nominal", color="#8FAADC")
ax.bar([i + w/2 for i in x], real_vals, width=w, label="Rendement REEL (Fisher)", color="#1F3864")

ax.axhline(rf_reel * 100, color="#C00000", linestyle="--", linewidth=1.5,
           label=f"Seuil Rf reel = {rf_reel*100:.2f}%")
ax.axhline(0, color="black", linewidth=0.8)

ax.set_xticks(list(x))
ax.set_xticklabels([f"{t}\n{names[t]}" for t in tickers_sorted], fontsize=8)
ax.set_ylabel("Rendement sur la periode (%)")
ax.set_title("BAA : rendement nominal vs rendement reel net d'inflation", fontsize=12)
ax.legend(fontsize=9)
ax.spines[["top", "right"]].set_visible(False)

for i, (nv, rv) in enumerate(zip(nominal_vals, real_vals)):
    ax.text(i - w/2, nv + (1.5 if nv >= 0 else -3), f"{nv:.1f}", ha="center", fontsize=7)
    ax.text(i + w/2, rv + (1.5 if rv >= 0 else -3), f"{rv:.1f}", ha="center", fontsize=7)

plt.tight_layout()
plt.savefig("02_rendement_reel_inflation.png", dpi=150)
print("\nGraphique sauvegarde : 02_rendement_reel_inflation.png")
