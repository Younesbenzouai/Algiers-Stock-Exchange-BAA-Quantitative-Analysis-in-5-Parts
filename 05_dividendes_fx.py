"""
Topic LinkedIn #5 -- "Dividendes et risque de change : le cas BIOPHARM"
Bourse d'Alger (BAA) : rendement des dividendes, total return et sensibilite EUR/DZD.

Ce script lit directement le classeur Excel BAA_Prix_Dividendes_Portefeuille_v6.xlsx
(feuilles "Statistiques", "Capitalisation" et "Taux de Change") au lieu de donnees codees en dur.

NOTE : l'exposition qualitative au risque de change (0=faible/1=moderee/2=forte) est une
evaluation d'analyste basee sur le secteur (import de matieres premieres), pas une donnee
du classeur -- modifiable dans SECTOR_FX_EXPOSURE ci-dessous.

Necessite : openpyxl, matplotlib
Auteur : Younes Benzouai
"""

import sys
import os
import openpyxl
import matplotlib.pyplot as plt

DEFAULT_EXCEL_PATH = r"C:\Users\youne\OneDrive\Desktop\LinkedIn\DZ\Outil d'optimisation de portefeuille boursier\BAA_Prix_Dividendes_Portefeuille_v6.xlsx"


def _resolve_excel_path(default_path):
    """Priorite : argument CLI > variable d'env BAA_XLSX_PATH > fichier .xlsx present
    a cote de ce script (cas du depot GitHub) > chemin par defaut. Ignore les arguments
    injectes par Jupyter/IPython (ex: '-f ...kernel.json')."""
    is_notebook = "ipykernel" in sys.argv[0] or "ipykernel" in sys.modules
    if not is_notebook and len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        return sys.argv[1]
    env_path = os.environ.get("BAA_XLSX_PATH")
    if env_path:
        return env_path
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        script_dir = os.getcwd()
    local_path = os.path.join(script_dir, "BAA_Prix_Dividendes_Portefeuille_v6.xlsx")
    if os.path.exists(local_path):
        return local_path
    return default_path


EXCEL_PATH = _resolve_excel_path(DEFAULT_EXCEL_PATH)

SECTOR_FX_EXPOSURE = {
    "Pharma": 2, "Hôtellerie": 1, "Investissement": 1,
    "Banque": 0, "Assurances": 0, "Finance": 0,
}


def _open(path):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fichier introuvable : {path}\n"
            "-> Verifie DEFAULT_EXCEL_PATH en haut du script, ou definis la variable\n"
            "   d'environnement BAA_XLSX_PATH avant de lancer le script/la cellule."
        )
    return openpyxl.load_workbook(path, data_only=True)


def load_statistiques(path, sheet_name="Statistiques"):
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Indicateur":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError("Ligne d'en-tete 'Indicateur' introuvable dans la feuille Statistiques.")

    header_cells = ws[header_row]
    tickers, names = [], {}
    for c in header_cells[1:]:
        if c.value is None:
            break
        parts = str(c.value).split("\n")
        t = parts[0].strip()
        tickers.append(t)
        names[t] = parts[1].strip() if len(parts) > 1 else t

    indicators = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        label = row[0].value
        if not label or not isinstance(label, str):
            continue
        values = [c.value for c in row[1:1 + len(tickers)]]
        if all(v is None for v in values):
            continue
        indicators[label.strip()] = dict(zip(tickers, values))

    return tickers, names, indicators


def load_sectors(path, tickers, sheet_name="Capitalisation"):
    """Recupere le secteur de chaque ticker depuis la feuille Capitalisation."""
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Ticker":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Ligne d'en-tete 'Ticker' introuvable dans la feuille {sheet_name}.")

    headers = [c.value for c in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    sectors = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        ticker = row[col["Ticker"]].value
        if ticker is None:
            break
        if str(ticker).strip().upper().startswith("TOTAL"):
            break
        sectors[str(ticker).strip()] = row[col["Secteur"]].value

    return {t: sectors.get(t, "Autre") for t in tickers}


def load_eur_dzd_variation(path, sheet_name="Taux de Change"):
    """Lit la variation EUR/DZD sur la periode d'analyse (tableau resume des paires cles)."""
    wb = _open(path)
    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Paire":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Ligne d'en-tete 'Paire' introuvable dans la feuille {sheet_name}.")

    headers = [c.value for c in ws[header_row]]
    col = {name: idx for idx, name in enumerate(headers) if name}

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        paire = row[0].value
        if paire is None or "/DZD" not in str(paire):
            break
        if str(paire).strip() == "EUR/DZD":
            return float(row[col["Variation (%)"]].value)

    raise ValueError("Paire EUR/DZD introuvable dans le tableau resume des taux de change.")


tickers, names, stats = load_statistiques(EXCEL_PATH)
sectors = load_sectors(EXCEL_PATH, tickers)
eur_dzd_var = load_eur_dzd_variation(EXCEL_PATH)

price_change = stats["Variation Prix (%)"]
div_yield = stats["Rendement Dividendes (%)"]
total_return = {t: (price_change[t] or 0) + (div_yield[t] or 0) for t in tickers}
fx_exposure = {t: SECTOR_FX_EXPOSURE.get(sectors[t], 0) for t in tickers}

print(f"Source : {EXCEL_PATH}")
print(f"Variation EUR/DZD sur la periode : {eur_dzd_var*100:+.1f}% (depreciation du dinar si positif)\n")
print(f"{'Ticker':6s} {'Div. yield':>11s} {'Var. prix':>11s} {'Total return':>13s} {'Exposition FX':>14s}")
for t in sorted(tickers, key=lambda x: -(div_yield[x] or 0)):
    exp_label = {0: "faible", 1: "moderee", 2: "FORTE (import)"}[fx_exposure[t]]
    print(f"{t:6s} {(div_yield[t] or 0)*100:10.2f}% {(price_change[t] or 0)*100:10.2f}% {total_return[t]*100:12.2f}% {exp_label:>14s}")

best = max(tickers, key=lambda t: div_yield[t] or 0)
print(f"\n{best} ({names[best]}) : meilleur rendement dividende de la cote.")
if fx_exposure[best] == 2:
    print(f"Exposition FX forte (secteur {sectors[best]}) : la depreciation du dinar (+{eur_dzd_var*100:.1f}% EUR/DZD)")
    print("constitue un vent contraire sur ses marges futures, non reflete dans la performance boursiere passee.")

tickers_sorted = sorted(tickers, key=lambda x: -(div_yield[x] or 0))
divs = [(div_yield[t] or 0) * 100 for t in tickers_sorted]
prices = [(price_change[t] or 0) * 100 for t in tickers_sorted]

fig, ax = plt.subplots(figsize=(11, 6.5))
x = range(len(tickers_sorted))

ax.bar(x, divs, label="Rendement dividende (%)", color="#2E7D32")
ax.bar(x, prices, label="Variation de prix (%)", color="#8FAADC",
       bottom=[d if p >= 0 else 0 for d, p in zip(divs, prices)])

for i, t in enumerate(tickers_sorted):
    tot = total_return[t] * 100
    y_top = max(divs[i] + max(prices[i], 0), 0.5)
    ax.text(i, y_top + 1.5, f"Total: {tot:+.1f}%", ha="center", fontsize=8, fontweight="bold")
    if fx_exposure[t] == 2:
        ax.text(i, -8, "⚠ FX", ha="center", fontsize=8, color="#C00000", fontweight="bold")

ax.axhline(0, color="black", linewidth=0.8)
ax.set_xticks(list(x))
ax.set_xticklabels([f"{t}\n{names[t]}" for t in tickers_sorted], fontsize=8)
ax.set_ylabel("% sur la periode analysee")
ax.set_title("BAA : rendement dividende vs variation de prix\n(marque ⚠ = forte exposition risque de change import)", fontsize=12)
ax.legend(["Rendement dividende (%)", "Variation de prix (%)"], fontsize=9, loc="upper right")
ax.spines[["top", "right"]].set_visible(False)

plt.tight_layout()
plt.savefig("05_dividendes_fx.png", dpi=150)
print("\nGraphique sauvegarde : 05_dividendes_fx.png")
